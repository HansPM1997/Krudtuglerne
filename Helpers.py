from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

def restore_left_border_col_E(ws):
    col_e = 5
    end_row = 500  # or max(ws.max_row, 500) if you want to be safe

    left_side = Side(style="thin")

    for r in range(1, end_row + 1):
        cell = ws.cell(row=r, column=col_e)
        old_border = cell.border

        cell.border = Border(
            left=left_side,
            right=old_border.right,
            top=old_border.top,
            bottom=old_border.bottom,
            diagonal=old_border.diagonal,
            diagonal_direction=old_border.diagonal_direction,
            outline=old_border.outline,
            vertical=old_border.vertical,
            horizontal=old_border.horizontal,
        )

def autofit_cell(ws, cell):
    text = str(cell.value) if cell.value is not None else ""
    col_letter = get_column_letter(cell.column)
    ws.column_dimensions[col_letter].width = len(text) + 2