from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from io import BytesIO

import Helpers as Help

def setup_participants_and_rounds(
    tema: str,
    lokation: str,
    dato: str,
    deltagere: list,
    antal_runder: int,
):
    # Project directory = folder containing this script
    project_dir = Path(__file__).resolve().parent

    filename_in = project_dir / "Template.xlsx"

    wb = load_workbook(filename_in)

    ws = wb['Ark1']

    # Write Navn, Lokation, Dato
    ws["C3"] = tema
    ws["C5"] = lokation
    ws["C7"] = str(dato)

    # Insert names
    start_row = 8
    start_col = 5
    num_of_participants = len(deltagere)

    for i in range(num_of_participants):
        cell = ws.cell(row=start_row, column=start_col + i)
        cell.value = deltagere[i]
        # make it bold
        cell.font = Font(bold=True)

    # After the loop: place "Gennemsnit" in the next column
    gennemsnit_col = start_col + num_of_participants

    g_cell = ws.cell(row=start_row, column=gennemsnit_col)
    g_cell.value = "Gennemsnit"
    g_cell.font = Font(bold=True)
    Help.autofit_cell(ws, g_cell)

    # Apply left border from top of sheet to bottom in the "Gennemsnit" column
    end_row = max(ws.max_row, 500)

    for r in range(1, end_row + 1):
        c = ws.cell(row=r, column=gennemsnit_col)
        c.border = Border(
             left=Side(style="thin"),
             right=c.border.right,
             top=c.border.top,
             bottom=c.border.bottom,
             diagonal=c.border.diagonal,
             diagonal_direction=c.border.diagonal_direction,
             outline=c.border.outline,
             vertical=c.border.vertical,
             horizontal=c.border.horizontal,
         )

    # Insert averages
    rounds_start_row = 10  # first row with scores
    first_score_col = start_col  # E
    last_score_col = gennemsnit_col - 1  # K (column before Gennemsnit)

    first_score_letter = get_column_letter(first_score_col)  # "E"
    last_score_letter = get_column_letter(last_score_col)  # "K"
    gennemsnit_col_letter = get_column_letter(gennemsnit_col)  # "L"

    # column to the right of all AVERAGE entries (VAR.P column)
    var_col = gennemsnit_col + 1
    var_col_letter = get_column_letter(var_col)  # e.g. "M"
    ws[f"{var_col_letter}8"] = "Varians"
    ws[f"{var_col_letter}8"].font = Font(bold=True)

    for p in range(num_of_participants):
        # top row of this participant's block
        base_row = rounds_start_row + p * (antal_runder + 1)

        for r in range(antal_runder):
            row = base_row + r  # actual row for this round

            # L[row] = AVERAGE(Erow:Krow)
            avg_formula = f"=AVERAGE({first_score_letter}{row}:{last_score_letter}{row})"
            avg_cell = ws[f"{gennemsnit_col_letter}{row}"]
            avg_cell.value = avg_formula

            # M[row] = VAR.P(Erow:Krow)  (to the right of AVERAGE)
            var_formula = f"=VAR.P({first_score_letter}{row}:{last_score_letter}{row})"
            var_cell = ws[f"{var_col_letter}{row}"]
            var_cell.value = var_formula

    # Fill in participant names
    header_start_row = 9  # first grey row
    header_start_col = 1  # column A
    last_fill_col = 20  # column T

    grey_fill = PatternFill(
        start_color="D9D9D9",
        end_color="D9D9D9",
        fill_type="solid",
    )

    for p in range(num_of_participants):
        row = header_start_row + p * (antal_runder + 1)

        for col in range(header_start_col, last_fill_col + 1):
            cell = ws.cell(row=row, column=col)

            # save existing border
            old_border = cell.border.copy()

            # apply fill
            cell.fill = grey_fill

            # restore border (including the left border in column E)
            cell.border = old_border

        # participant name in column B
        name_cell = ws.cell(row=row, column=2)
        name_cell.value = deltagere[p]
        name_cell.font = Font(bold=True)

    # Restore border in column E
    Help.restore_left_border_col_E(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


